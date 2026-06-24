import os
import sys
import json
import argparse
import subprocess
from PIL import Image

# Third-party imports that we installed
try:
    from pypdf import PdfWriter, PdfReader, PageObject, Transformation
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
except ImportError:
    # If the libraries are not available yet (still installing), we'll defer errors to execution time
    pass

# Try to register standard macOS / Windows fonts for Chinese rendering support
DEFAULT_FONT = 'Helvetica'
try:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    # Candidate font paths for macOS and Windows
    font_candidates = [
        "/System/Library/Fonts/STHeiti Light.ttc",  # Mac Heiti
        "/System/Library/Fonts/PingFang.ttc",       # Mac PingFang
        "C:\\Windows\\Fonts\\msyh.ttc",              # Windows Microsoft YaHei
        "C:\\Windows\\Fonts\\simsun.ttc",            # Windows SimSun
    ]
    
    font_path = None
    for path in font_candidates:
        if os.path.exists(path):
            font_path = path
            break
            
    if font_path:
        pdfmetrics.registerFont(TTFont('Heiti', font_path))
        DEFAULT_FONT = 'Heiti'
except Exception as e:
    pass

def unzip_file(file_path, out_dir):
    """
    Extract zip or rar file.
    Rar files are extracted using 'unar' CLI tool.
    Zip files use Python's zipfile or unar.
    """
    os.makedirs(out_dir, exist_ok=True)
    file_ext = os.path.splitext(file_path)[1].lower()
    
    if file_ext == '.zip':
        import zipfile
        try:
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                zip_ref.extractall(out_dir)
            return {"status": "success", "message": "Extracted zip successfully"}
        except Exception as e:
            return {"status": "error", "message": f"Failed to extract zip via Python: {str(e)}"}
            
    elif file_ext in ['.rar', '.cbr']:
        # Try using unar
        try:
            # unar -o <out_dir> -f <file_path> (-f is force overwrite)
            result = subprocess.run(
                ["unar", "-o", out_dir, "-f", file_path],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )
            if result.returncode == 0:
                return {"status": "success", "message": "Extracted rar successfully via unar"}
            else:
                return {"status": "error", "message": f"unar exited with code {result.returncode}: {result.stderr}"}
        except FileNotFoundError:
            return {
                "status": "error", 
                "message": "unar command not found. Please install it using 'brew install unar' to extract RAR files."
            }
    else:
        return {"status": "error", "message": f"Unsupported archive format: {file_ext}"}

def scan_directory(dir_path):
    """
    Scan directory structure. Detect vehicle folders (VIN folders) and other documents.
    """
    if not os.path.exists(dir_path):
        return {"status": "error", "message": f"Directory {dir_path} does not exist"}
        
    result = {
        "root_dir": dir_path,
        "vehicle_folders": [],
        "other_files": []
    }
    
    # We walk the directory.
    # Typically WeChat unzipped files have deep structures: e.g. PackName/PackName/Subdirs...
    # We find folders that look like vehicle VIN folders or category folders.
    # Usually: '1 买卖合同和合格证' contains subfolders like '1-LS6CME...', '2-LS5A2D...'
    
    for root, dirs, files in os.walk(dir_path):
        # Filter hidden files
        files = [f for f in files if not f.startswith('.')]
        dirs[:] = [d for d in dirs if not d.startswith('.')]
        
        # If a directory name has a dash and looks like a vehicle VIN/code (e.g. "1-LS6CME...")
        # or it is inside '1 买卖合同和合格证' and contains images/pdfs:
        # We classify it as a vehicle folder if it has parent '1 买卖合同和合格证'
        parent_name = os.path.basename(root)
        
        # Check if the folder itself contains image or pdf files, and looks like a vehicle entry
        is_vehicle_folder = False
        if '-' in parent_name and any(parent_name.split('-')[0].isdigit() for part in parent_name.split('-')):
            # e.g., "1-LS6C..." or similar
            is_vehicle_folder = True
            
        if is_vehicle_folder:
            vehicle_entry = {
                "name": parent_name,
                "path": root,
                "files": []
            }
            for f in files:
                f_path = os.path.join(root, f)
                f_ext = os.path.splitext(f)[1].lower()
                f_type = "unknown"
                if f_ext in ['.jpg', '.jpeg', '.png']:
                    f_type = "image"
                elif f_ext == '.pdf':
                    f_type = "pdf"
                elif f_ext in ['.xlsx', '.xls']:
                    f_type = "excel"
                elif f_ext in ['.docx', '.doc']:
                    f_type = "word"
                    
                vehicle_entry["files"].append({
                    "name": f,
                    "path": f_path,
                    "type": f_type,
                    "size_bytes": os.path.getsize(f_path)
                })
            # Sort files so images and PDFs are in predictable order (e.g., pdf first then image)
            vehicle_entry["files"].sort(key=lambda x: (x["type"] != "pdf", x["name"]))
            result["vehicle_folders"].append(vehicle_entry)
            
        else:
            # Check files in root or other folders (like root of unzipped files)
            # If they are not in a vehicle folder, we treat them as 'other_files'
            # But we only want files directly in the root or main folders, not inside vehicle subfolders
            # We determine if we are inside a vehicle folder path
            in_vehicle_path = False
            for v_folder in result["vehicle_folders"]:
                if root.startswith(v_folder["path"]):
                    in_vehicle_path = True
                    break
                    
            if not in_vehicle_path:
                for f in files:
                    f_path = os.path.join(root, f)
                    f_ext = os.path.splitext(f)[1].lower()
                    
                    # 优先检测是否有同名 PDF 副本
                    if f_ext in ['.xlsx', '.xls', '.docx', '.doc']:
                        base_name_no_ext = os.path.splitext(f)[0]
                        pdf_sibling = os.path.join(root, base_name_no_ext + ".pdf")
                        if os.path.exists(pdf_sibling):
                            f = base_name_no_ext + ".pdf"
                            f_path = pdf_sibling
                            f_ext = '.pdf'
                            
                    f_type = "unknown"
                    if f_ext in ['.jpg', '.jpeg', '.png']:
                        f_type = "image"
                    elif f_ext == '.pdf':
                        f_type = "pdf"
                    elif f_ext in ['.xlsx', '.xls']:
                        f_type = "excel"
                    elif f_ext in ['.docx', '.doc']:
                        f_type = "word"
                        
                    result["other_files"].append({
                        "name": f,
                        "path": f_path,
                        "type": f_type,
                        "size_bytes": os.path.getsize(f_path)
                    })

    # Sort vehicle folders by their prefix number if possible
    def get_folder_sort_key(folder):
        name = folder["name"]
        parts = name.split('-')
        if parts and parts[0].isdigit():
            return (int(parts[0]), name)
        return (9999, name)
        
    result["vehicle_folders"].sort(key=get_folder_sort_key)
    
    return {"status": "success", "data": result}

def image_to_pdf(image_path, output_pdf_path):
    """
    Convert an image to an A4 PDF page, keeping aspect ratio and fitting inside margins (anti-overflow).
    Uses PIL to pre-decode and convert to standard RGB mode for perfect PDF renderer compatibility.
    """
    temp_std_img = None
    try:
        img = Image.open(image_path)
        if img.mode != 'RGB':
            img = img.convert('RGB')
        w_orig, h_orig = img.size
        
        # A4 standard size is 595.27 x 841.89 points
        page_w, page_h = A4
        
        # 15 points margin
        margin = 15
        avail_w = page_w - (margin * 2)
        avail_h = page_h - (margin * 2)
        
        # Calculate scale factor
        scale_w = avail_w / w_orig
        scale_h = avail_h / h_orig
        scale = min(scale_w, scale_h)
        
        # Determine final width and height on A4 (never upscale beyond original size or A4 bounds)
        w_new = w_orig * min(scale, 1.0)
        h_new = h_orig * min(scale, 1.0)
        
        # Center the image
        x = (page_w - w_new) / 2
        y = (page_h - h_new) / 2
        
        # Save a temporary clean JPEG file that ReportLab can easily read
        temp_std_img = output_pdf_path + "_temp_clean.jpg"
        img.save(temp_std_img, "JPEG", quality=95)
        
        c = canvas.Canvas(output_pdf_path, pagesize=A4)
        c.drawImage(temp_std_img, x, y, width=w_new, height=h_new)
        c.showPage()
        c.save()
        return True
    except Exception as e:
        print(f"Error converting image {image_path}: {e}", file=sys.stderr)
        return False
    finally:
        # Cleanup temporary image file
        if temp_std_img and os.path.exists(temp_std_img):
            try:
                os.remove(temp_std_img)
            except:
                pass

def fit_pdf_to_single_page(pdf_path):
    reader = PdfReader(pdf_path)
    if len(reader.pages) <= 1:
        return

    page_w, page_h = A4
    margin = 12
    slot_h = (page_h - margin * 2) / len(reader.pages)
    merged_page = PageObject.create_blank_page(width=page_w, height=page_h)

    # ponytail: single-sheet Excel must not paginate; if too dense, it shrinks.
    for idx, page in enumerate(reader.pages):
        src_w = float(page.mediabox.width)
        src_h = float(page.mediabox.height)
        if src_w <= 0 or src_h <= 0:
            continue

        scale = min((page_w - margin * 2) / src_w, slot_h / src_h)
        x = (page_w - src_w * scale) / 2
        y = page_h - margin - (idx + 1) * slot_h + (slot_h - src_h * scale) / 2
        merged_page.merge_transformed_page(page, Transformation().scale(scale).translate(x, y))

    temp_path = pdf_path + ".single-page.pdf"
    writer = PdfWriter()
    writer.add_page(merged_page)
    with open(temp_path, "wb") as f:
        writer.write(f)
    os.replace(temp_path, pdf_path)

def excel_sheet_count(excel_path):
    file_ext = os.path.splitext(excel_path)[1].lower()
    if file_ext == '.xlsx':
        import openpyxl
        return len(openpyxl.load_workbook(excel_path, read_only=True).sheetnames)
    if file_ext == '.xls':
        import xlrd
        return xlrd.open_workbook(excel_path, on_demand=True).nsheets
    return 0

def prepare_office_pdf(pdf_path, temp_pdfs, force_single_page=False):
    if force_single_page:
        fit_pdf_to_single_page(pdf_path)
    temp_pdfs.append(pdf_path)
    return pdf_path

def convert_office_to_pdf(office_path, output_dir, temp_pdfs):
    """
    Convert xlsx/xls/docx/doc to PDF.
    Tries LibreOffice first. If not present, falls back to simplified pure Python conversions.
    """
    file_ext = os.path.splitext(office_path)[1].lower()
    base_name = os.path.splitext(os.path.basename(office_path))[0]
    output_pdf_name = f"{base_name}.pdf"
    output_pdf_path = os.path.join(output_dir, output_pdf_name)
    single_sheet_excel = file_ext in ['.xlsx', '.xls'] and excel_sheet_count(office_path) == 1
    
    # 1. Try LibreOffice soffice
    libreoffice_paths = [
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",  # macOS standard
        "soffice"  # PATH
    ]
    
    soffice_exe = None
    for path in libreoffice_paths:
        if path.startswith("/") and os.path.exists(path):
            soffice_exe = path
            break
        elif not path.startswith("/"):
            try:
                # Check if soffice is in PATH
                subprocess.run([path, "--version"], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                soffice_exe = path
                break
            except FileNotFoundError:
                pass
                
    if soffice_exe:
        try:
            # soffice --headless --convert-to pdf --outdir <output_dir> <office_path>
            convert_to = "pdf"
            if file_ext in ['.xlsx', '.xls']:
                convert_to = 'pdf:calc_pdf_Export:{"SinglePageSheets":{"type":"boolean","value":"true"}}'
            result = subprocess.run(
                [soffice_exe, "--headless", "--convert-to", convert_to, "--outdir", output_dir, office_path],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )
            if result.returncode == 0 and os.path.exists(output_pdf_path):
                return prepare_office_pdf(output_pdf_path, temp_pdfs, single_sheet_excel)
        except Exception as e:
            print(f"LibreOffice conversion failed: {e}", file=sys.stderr)
            
    # 2. Fallback to native python for Excel
    if file_ext in ['.xlsx', '.xls']:
        try:
            success = excel_to_pdf_fallback(office_path, output_pdf_path)
            if success and os.path.exists(output_pdf_path):
                return prepare_office_pdf(output_pdf_path, temp_pdfs, single_sheet_excel)
        except Exception as e:
            print(f"Excel fallback conversion failed: {e}", file=sys.stderr)
            
    # 3. Fallback to native python for Word (.docx)
    elif file_ext == '.docx':
        try:
            success = docx_to_pdf_fallback(office_path, output_pdf_path)
            if success and os.path.exists(output_pdf_path):
                return prepare_office_pdf(output_pdf_path, temp_pdfs)
        except Exception as e:
            print(f"Word fallback conversion failed: {e}", file=sys.stderr)
            
    # 4. Fallback: Write error/info page
    try:
        write_error_pdf_page(output_pdf_path, f"Document: {os.path.basename(office_path)}\n\n"
                                              f"Notice: This file could not be converted. Please install LibreOffice\n"
                                              f"(brew install --cask libreoffice) on your Mac for perfect Office-to-PDF conversion.")
        temp_pdfs.append(output_pdf_path)
        return output_pdf_path
    except Exception as e:
        print(f"Failed to generate notice page: {e}", file=sys.stderr)
        return None

def excel_to_pdf_fallback(excel_path, output_pdf_path):
    """
    Simplified Excel to PDF conversion fallback using openpyxl/xlrd & reportlab.
    """
    import html
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    file_ext = os.path.splitext(excel_path)[1].lower()
    sheets_data = []
    
    if file_ext == '.xlsx':
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        for sheet in wb.worksheets:
            data = []
            max_r = min(sheet.max_row + 1, 60)      # Cap at 60 rows
            max_c = min(sheet.max_column + 1, 10)   # Cap at 10 columns
            for r in range(1, max_r):
                row_data = []
                for c in range(1, max_c):
                    val = sheet.cell(row=r, column=c).value
                    row_data.append(str(val) if val is not None else "")
                if any(row_data):
                    data.append(row_data)
            if data:
                sheets_data.append(data)
    elif file_ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(excel_path)
        for sheet in wb.sheets():
            data = []
            max_r = min(sheet.nrows, 60)
            max_c = min(sheet.ncols, 10)
            for r in range(max_r):
                row_data = []
                for c in range(max_c):
                    val = sheet.cell_value(rowx=r, colx=c)
                    if isinstance(val, float) and val.is_integer():
                        val = int(val)
                    row_data.append(str(val) if val != "" else "")
                if any(row_data):
                    data.append(row_data)
            if data:
                sheets_data.append(data)
    else:
        raise ValueError(f"Unsupported spreadsheet format: {file_ext}")

    doc = SimpleDocTemplate(output_pdf_path, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    story = []
    
    styles = getSampleStyleSheet()
    
    if sheets_data:
        avail_w = A4[0] - 40
        cell_style = ParagraphStyle(
            'ExcelCellChinese',
            parent=styles['Normal'],
            fontName=DEFAULT_FONT,
            fontSize=6,
            leading=8,
            wordWrap='CJK'
        )
        for sheet_idx, data in enumerate(sheets_data):
            if sheet_idx:
                story.append(PageBreak())
            col_count = max(len(row) for row in data)
            weights = [
                min(max(max(len(row[c]) if c < len(row) else 0 for row in data), 4), 24)
                for c in range(col_count)
            ]
            col_widths = [avail_w * weight / sum(weights) for weight in weights]
            table_data = [
                [Paragraph(html.escape(row[c] if c < len(row) else ""), cell_style) for c in range(col_count)]
                for row in data
            ]
            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f2f2f2')),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
                ('FONTNAME', (0,0), (-1,-1), DEFAULT_FONT),
                ('FONTSIZE', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('TOPPADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(t)
    else:
        empty_style = ParagraphStyle(
            'ExcelEmptyChinese',
            parent=styles['Normal'],
            fontName=DEFAULT_FONT,
            fontSize=10
        )
        story.append(Paragraph("Empty spreadsheet sheet.", empty_style))
        
    doc.build(story)
    return True

def docx_to_pdf_fallback(docx_path, output_pdf_path):
    """
    Fallback Word to PDF parser using python-docx and reportlab.
    Preserves text paragraphs and tables in their original document order.
    """
    import docx
    from docx.text.paragraph import Paragraph as DocxParagraph
    from docx.table import Table as DocxTable
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    doc = docx.Document(docx_path)
    
    pdf_doc = SimpleDocTemplate(output_pdf_path, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    styles = getSampleStyleSheet()
    
    body_style = ParagraphStyle(
        'ChineseBody',
        parent=styles['Normal'],
        fontName=DEFAULT_FONT,
        fontSize=10,
        leading=14,
        spaceAfter=6
    )
    
    table_cell_style = ParagraphStyle(
        'TableCellChinese',
        parent=styles['Normal'],
        fontName=DEFAULT_FONT,
        fontSize=8,
        leading=10
    )
    
    # 遍历 body 里的子元素以保持物理顺序
    for child in doc.element.body:
        if child.tag.endswith('p'):
            p = DocxParagraph(child, doc)
            text = p.text.strip()
            if text:
                story.append(Paragraph(text, body_style))
        elif child.tag.endswith('tbl'):
            t = DocxTable(child, doc)
            # 提取表格数据
            table_data = []
            for row in t.rows:
                row_data = []
                for cell in row.cells:
                    cell_text = cell.text.strip()
                    row_data.append(Paragraph(cell_text, table_cell_style))
                table_data.append(row_data)
                
            if table_data:
                col_count = len(table_data[0])
                # A4 可用宽度为 595.27 - 60 = 535.27
                col_width = 535.27 / max(col_count, 1)
                
                table_flow = Table(table_data, colWidths=[col_width] * col_count)
                table_flow.setStyle(TableStyle([
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#999999')),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                    ('TOPPADDING', (0,0), (-1,-1), 4),
                    ('LEFTPADDING', (0,0), (-1,-1), 4),
                    ('RIGHTPADDING', (0,0), (-1,-1), 4),
                ]))
                story.append(table_flow)
                story.append(Spacer(1, 10))
                
    if not story:
        story.append(Paragraph("Empty Word document.", body_style))
        
    pdf_doc.build(story)
    return True

def write_error_pdf_page(pdf_path, text):
    """
    Generate a simple PDF page showing a message.
    """
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)
    styles = getSampleStyleSheet()
    
    heading_style = ParagraphStyle(
        'NoticeHeading',
        parent=styles['Heading1'],
        fontName=DEFAULT_FONT,
        fontSize=18,
        leading=22
    )
    
    body_style = ParagraphStyle(
        'NoticeBody',
        parent=styles['Normal'],
        fontName=DEFAULT_FONT,
        fontSize=10,
        leading=14
    )
    
    story = []
    story.append(Paragraph("Document Conversion Notice", heading_style))
    story.append(Spacer(1, 20))
    
    for line in text.split('\n'):
        if line.strip():
            story.append(Paragraph(line, body_style))
            story.append(Spacer(1, 8))
            
    doc.build(story)

def merge_pdf_files(files_json_path, output_pdf_path):
    """
    Merge a list of files (PDF, images, Office documents) in specified order.
    The order and details are read from a JSON file.
    """
    if not os.path.exists(files_json_path):
        return {"status": "error", "message": f"Input JSON file {files_json_path} does not exist"}
        
    try:
        with open(files_json_path, 'r', encoding='utf-8') as f:
            file_entries = json.load(f)
    except Exception as e:
        return {"status": "error", "message": f"Failed to parse input JSON: {e}"}
        
    merger = PdfWriter()
    temp_pdfs = []
    output_dir = os.path.dirname(output_pdf_path)
    os.makedirs(output_dir, exist_ok=True)
    
    # We process each file:
    # 1. Images: Convert to A4 PDF first
    # 2. Excel/Word: Convert to PDF
    # 3. PDF: Append directly
    
    processed_count = 0
    try:
        for idx, entry in enumerate(file_entries):
            f_path = entry.get("path")
            f_type = entry.get("type")
            
            # 双重保险：如果有同名 PDF，直接用 PDF 替代合并
            if f_type in ["excel", "word"] and f_path:
                base_path_no_ext = os.path.splitext(f_path)[0]
                pdf_sibling = base_path_no_ext + ".pdf"
                if os.path.exists(pdf_sibling):
                    f_path = pdf_sibling
                    f_type = "pdf"
            
            if not f_path or not os.path.exists(f_path):
                print(f"Skipping missing file: {f_path}", file=sys.stderr)
                continue
                
            if f_type == "pdf":
                # Verify PDF is valid before appending
                try:
                    with open(f_path, 'rb') as pdf_file:
                        reader = PdfReader(pdf_file)
                        _ = len(reader.pages) # Touch pages to verify
                    merger.append(f_path)
                    processed_count += 1
                except Exception as e:
                    print(f"Skipping corrupted PDF {f_path}: {e}", file=sys.stderr)
                    
            elif f_type == "image":
                temp_pdf = os.path.join(output_dir, f"_temp_img_{idx}.pdf")
                success = image_to_pdf(f_path, temp_pdf)
                if success:
                    temp_pdfs.append(temp_pdf)
                    merger.append(temp_pdf)
                    processed_count += 1
                    
            elif f_type in ["excel", "word"]:
                pdf_path = convert_office_to_pdf(f_path, output_dir, temp_pdfs)
                if pdf_path and os.path.exists(pdf_path):
                    merger.append(pdf_path)
                    processed_count += 1
            else:
                print(f"Skipping unknown file type '{f_type}': {f_path}", file=sys.stderr)
                
        if processed_count > 0:
            merger.write(output_pdf_path)
            merger.close()
            
            # Cleanup temporary files
            for temp_pdf in temp_pdfs:
                try:
                    os.remove(temp_pdf)
                except:
                    pass
                    
            return {"status": "success", "message": f"Successfully merged {processed_count} files into {output_pdf_path}"}
        else:
            merger.close()
            return {"status": "error", "message": "No valid files were processed or merged."}
            
    except Exception as e:
        merger.close()
        # Attempt cleanup on error
        for temp_pdf in temp_pdfs:
            try:
                os.remove(temp_pdf)
            except:
                pass
        return {"status": "error", "message": f"Merge failed: {e}"}

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="FDR Document Processor Helper")
    parser.add_argument("--action", required=True, choices=["unzip", "scan", "merge"])
    parser.add_argument("--file", help="Archive file path or merge JSON list file path")
    parser.add_argument("--outdir", help="Output extraction directory or temporary directory")
    parser.add_argument("--dir", help="Directory path to scan")
    parser.add_argument("--output", help="Output file path (e.g. output PDF)")
    
    args = parser.parse_args()
    
    if args.action == "unzip":
        if not args.file or not args.outdir:
            print(json.dumps({"status": "error", "message": "unzip action requires --file and --outdir"}))
            sys.exit(1)
        res = unzip_file(args.file, args.outdir)
        print(json.dumps(res, ensure_ascii=False))
        
    elif args.action == "scan":
        if not args.dir:
            print(json.dumps({"status": "error", "message": "scan action requires --dir"}))
            sys.exit(1)
        res = scan_directory(args.dir)
        print(json.dumps(res, ensure_ascii=False))
        
    elif args.action == "merge":
        if not args.file or not args.output:
            print(json.dumps({"status": "error", "message": "merge action requires --file (JSON config) and --output"}))
            sys.exit(1)
        res = merge_pdf_files(args.file, args.output)
        print(json.dumps(res, ensure_ascii=False))
